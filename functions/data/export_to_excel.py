"""Exporta los atributos de la capa activa a un archivo Excel"""

from qgis.core import QgsProject, QgsVectorLayer
from qgis.PyQt.QtWidgets import QFileDialog
import os
from datetime import datetime

def ejecutar(iface, params=None):
    """
    Exporta la tabla de atributos de la capa activa a Excel
    
    Args:
        iface: Interfaz de QGIS
        params: Parámetros opcionales (ruta de salida)
    
    Returns:
        dict: Estado y mensaje del resultado
    """
    # Verificar que openpyxl está disponible
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {
            "status": "error",
            "mensaje": "Necesita instalar openpyxl:\npip install openpyxl"
        }
    
    capa = iface.activeLayer()
    
    if not capa:
        return {
            "status": "warning",
            "mensaje": "Seleccione una capa para exportar"
        }
    
    if capa.type() != 0:  # 0 = Vector Layer
        return {
            "status": "error",
            "mensaje": "Solo se pueden exportar capas vectoriales"
        }
    
    try:
        # Determinar archivo de salida
        if params and 'output_path' in params:
            archivo_salida = params['output_path']
        else:
            # Usar diálogo para seleccionar ubicación
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_sugerido = f"{capa.name()}_{timestamp}.xlsx"
            archivo_salida = os.path.join(
                os.path.expanduser("~/Desktop"),
                nombre_sugerido
            )
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = capa.name()[:31]  # Excel limita a 31 caracteres
        
        # Obtener campos
        campos = capa.fields()
        nombres_campos = [campo.name() for campo in campos]
        
        # Escribir encabezados con formato
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", 
                                 end_color="366092", 
                                 fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, nombre_campo in enumerate(nombres_campos, 1):
            celda = ws.cell(row=1, column=col, value=nombre_campo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
        
        # Obtener entidades (seleccionadas o todas)
        if capa.selectedFeatureCount() > 0:
            entidades = capa.selectedFeatures()
            mensaje_seleccion = f"{capa.selectedFeatureCount()} entidades seleccionadas"
        else:
            entidades = capa.getFeatures()
            mensaje_seleccion = f"{capa.featureCount()} entidades"
        
        # Escribir datos
        fila = 2
        for entidad in entidades:
            for col, campo in enumerate(nombres_campos, 1):
                valor = entidad[campo]
                
                # Manejar tipos de datos especiales
                if valor is None:
                    valor = ""
                elif hasattr(valor, 'toString'):  # QDate, QDateTime
                    valor = valor.toString('yyyy-MM-dd')
                elif hasattr(valor, '__geo_interface__'):  # Geometría
                    valor = "GEOMETRY"
                
                ws.cell(row=fila, column=col, value=valor)
            
            fila += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Agregar filtros automáticos
        ws.auto_filter.ref = ws.dimensions
        
        # Guardar archivo
        wb.save(archivo_salida)
        
        # Abrir carpeta contenedora
        carpeta = os.path.dirname(archivo_salida)
        if os.name == 'nt':  # Windows
            os.startfile(carpeta)
        elif os.name == 'posix':  # macOS and Linux
            os.system(f'open "{carpeta}"' if os.uname().sysname == 'Darwin' 
                     else f'xdg-open "{carpeta}"')
        
        return {
            "status": "ok",
            "mensaje": f"Exportado exitosamente:\n{archivo_salida}\n({mensaje_seleccion})"
        }
        
    except PermissionError:
        return {
            "status": "error",
            "mensaje": "El archivo está abierto en otro programa. Ciérrelo e intente de nuevo."
        }
    except Exception as e:
        return {
            "status": "error",
            "mensaje": f"Error al exportar: {str(e)}"
        }
